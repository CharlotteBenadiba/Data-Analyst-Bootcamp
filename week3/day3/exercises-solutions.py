####### EXERCISE 1 : displaying rows 

# Import the pandas library
import pandas as pd

!pip install kaggle
from google.colab import files
files.upload()
!mkdir -p ~/.kaggle
!cp kaggle.json ~/.kaggle/
!chmod 600 ~/.kaggle/kaggle.json
!kaggle competitions download -c titanic
!unzip titanic.zip

# Path to the CSV file
csv_file_path = 'train.csv'  # Replace with the path to your CSV file

# Import the CSV file into a DataFrame
dataframe = pd.read_csv(csv_file_path)

# Display the first 
print(dataframe.head())


####### EXERCISE 2 : export df to excel and JSON

# Import the pandas library
import pandas as pd

# Create a simple DataFrame
data = {
    'Name': ['Alice', 'Bob', 'Charlie', 'David', 'Eva'],
    'Age': [25, 30, 35, 40, 45],
    'City': ['New York', 'Los Angeles', 'Chicago', 'Houston', 'Phoenix']
}

df = pd.DataFrame(data)

# Export the DataFrame to an Excel file
excel_filename = 'sample_data.xlsx'
df.to_excel(excel_filename, index=False)


print(f"DataFrame exported to {excel_filename}")

#export df to json
json_file_path = 'sample_data.json'  # Replace with your JSON file path
data_frame = pd.read_json(json_file_path)
print(data_frame)


####### EXERCISE 3 : export df to json with orient

# Import Pandas library
import pandas as pd

# Data Preparation
# Create a sample DataFrame
data = {
    'UserID': range(1, 11),
    'Name': ['Alice', 'Bob', 'Charlie', 'David', 'Emma', 
             'Frank', 'Grace', 'Hannah', 'Ian', 'Jane'],
    'Age': [25, 30, 35, 40, 29, 34, 28, 31, 36, 33],
    'City': ['New York', 'Los Angeles', 'Chicago', 'Houston', 'Phoenix',
             'Philadelphia', 'San Antonio', 'San Diego', 'Dallas', 'San Jose']
}
df = pd.DataFrame(data)

# Exporting DataFrame to JSON with different orientations
orientations = ['split', 'records', 'index', 'columns', 'values']

for orient in orientations:
    # Export to JSON
    df.to_json(f'data_{orient}.json', orient=orient)



####### EXERCISE 4 : export df from excel

  # Import necessary libraries
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Prepare the data
data = pd.DataFrame({
    'Date': pd.date_range('20210101', periods=20),
    'Product': ['Product A', 'Product B', 'Product C'] * 6 + ['Product A', 'Product B'],
    'Quantity': [2, 3, 5, 7, 1] * 4,
    'Revenue': [210.50, 340.60, 450.30, 120.10, 150.75] * 4
})

# Create a new Excel workbook and sheet
wb = Workbook()
ws = wb.active

# Adding data to the sheet
for r in dataframe_to_rows(data, index=False, header=True):
    ws.append(r)

# Apply formatting
# Header formatting
for cell in ws[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="4f81bd")
    cell.alignment = Alignment(horizontal="center")

# Set number format for Revenue and Quantity
for row in ws.iter_rows(min_row=2, max_col=4, max_row=21):
    row[2].number_format = '#,##0'  # Quantity column
    row[3].number_format = '#,##0.00'  # Revenue column

# Auto-adjust columns' width
for column in ws.columns:
    length = max(len(str(cell.value)) for cell in column)
    ws.column_dimensions[column[0].column_letter].width = length

# Save the workbook
wb.save('formatted_sales_report.xlsx')

print("DataFrame exported to 'formatted_sales_report.xlsx' with formatting.")

---

####### EXERCISE 5 : Chunking Large Datasets (medium)
# Import necessary library
import pandas as pd

# Define the path to the large dataset file
dtypes = {
    "row_id": "int64",
    "timestamp": "int64",
    "user_id": "int32",
    "content_id": "int16",
    "content_type_id": "boolean",
    "task_container_id": "int16",
    "user_answer": "int8",
    "answered_correctly": "int8",
    "prior_question_elapsed_time": "float32", 
    "prior_question_had_explanation": "boolean"
}

#print("Train size:", data.shape)

# Initialize a variable to store the sum of a specific column across chunks
# Assuming we're summing a column named 'column_of_interest'
sum_of_column = 0

# Read the dataset in chunks and process each chunk
chunk_size = 1000  # Define the chunk size
for chunk in pd.read_csv("../input/riiid-test-answer-prediction/train.csv", chunksize=chunk_size):
    # Perform a basic operation on the chunk
    # Example: calculate the sum of a specific column
    sum_of_column += chunk['column_of_interest'].sum()

    # Append chunk to an output file
    # In this example, we're just saving the chunk as is
    chunk.to_csv('output_chunk.csv', mode='a', header=False, index=False)

# Output the calculated sum
print(f"Sum of the column across all chunks: {sum_of_column}")

# Note: The script assumes that the 'column_of_interest' exists in the dataset
# Replace 'column_of_interest' with the actual column name you want to sum


####### EXERCISE 6 (medium)
# Import necessary library
import pandas as pd
import time

# Assuming large_df is the large DataFrame from the previous exercise
# For demonstration, let's create a large DataFrame (remove this if you have large_df already)
# large_df = pd.DataFrame({ ... })  # Replace with your DataFrame creation code

# Export to Parquet
# Start timer
start_time = time.time()
large_df.to_parquet('large_data.parquet')
end_time = time.time()

# Calculate time taken and print
time_taken = end_time - start_time
print(f"Time taken to write Parquet file: {time_taken} seconds")

# To demonstrate reading speeds, let's read the Parquet file back into a DataFrame
# Start timer
start_time = time.time()
parquet_df = pd.read_parquet('large_data.parquet')
end_time = time.time()

# Calculate time taken and print
time_taken = end_time - start_time
print(f"Time taken to read Parquet file: {time_taken} seconds")

# Optionally, compare file sizes on disk between the CSV and Parquet formats
# This part requires manual checking of the file sizes in the file system



####### EXERCISE 7 (hard) : Database Integration 

# Import necessary libraries
import pandas as pd
import sqlalchemy

# Set up the connection to the SQL Murder Mystery Database
# Download the database file from the provided Kaggle link and place it in an accessible directory
database_path = 'path/to/murder_mystery.db'  # Replace with the actual path
engine = sqlalchemy.create_engine(f'sqlite:///{database_path}')

# Assuming you have a merged DataFrame named 'merged_data'
# For demonstration, let's create a dummy DataFrame
merged_data = pd.DataFrame({'column1': [1, 2, 3], 'column2': ['A', 'B', 'C']})

# Export the merged DataFrame to the SQL database
merged_data.to_sql('merged_table', engine, if_exists='replace', index=False, method='multi')

# Perform a SQL query to retrieve a subset of the data
# This query can be customized based on the structure of 'merged_table'
query = 'SELECT * FROM merged_table WHERE column1 > 1'
subset = pd.read_sql_query(query, engine)

# Export this subset to an Excel file
subset.to_excel('subset_data.xlsx', index=False)

print("Data exported to SQL and subset exported to Excel.")


